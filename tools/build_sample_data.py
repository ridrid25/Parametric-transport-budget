#!/usr/bin/env python3
"""Собирает образцы данных (samples/) из out.inputs.json.

Порядок запуска:
    node tools/export_inputs.js      # создаёт tools/out.inputs.json
    python3 tools/build_sample_data.py

На выходе:
    samples/transport_budget_data.xlsx  — многовкладочный (Инструкция + входные данные)
    samples/transport_budget_data.csv   — те же данные одним листом с секциями
                                          (готово для Google Sheets)
"""
import json
import os
import io
import csv

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
INPUTS = os.path.join(HERE, 'out.inputs.json')
SAMPLES = os.path.join(ROOT, 'samples')


def build_csv(data):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['ТРАНСПОРТНЫЙ БЮДЖЕТ — ТЕСТОВЫЕ ДАННЫЕ И ПАРАМЕТРЫ'])
    w.writerow(['Полную базу ~1000 клиентов и бюджет генерирует скрипт apps-script/Code.gs.'])
    w.writerow(['Меняйте раздел ПАРАМЕТРЫ и запускайте меню «Транспортный бюджет» - «Собрать всё с нуля».'])
    w.writerow([])

    def section(title, rows):
        w.writerow([title])
        for r in rows:
            w.writerow(r)
        w.writerow([])

    section('ПАРАМЕТРЫ МОДЕЛИ (param / value / описание)', data['Params'])
    section('ТРАНСПОРТ — 13 машин', data['Vehicles'])
    section('СЕЗОННОСТЬ ПО МЕСЯЦАМ', data['Seasonality'])
    return buf.getvalue()


def build_xlsx(data, path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Инструкция'
    lines = [
        ('Транспортный бюджет — тестовые данные', Font(bold=True, size=13, color='1F4E79')),
        ('', None),
        ('Тиры транспорта, сезонность и параметры модели.', Font(bold=True)),
        ('Полную базу ~1000 клиентов и бюджет генерирует apps-script/Code.gs.', None),
        ('', None),
        ('Чтобы получить параметрическую модель (бюджет, сценарии, дашборд):', Font(bold=True)),
        ('  1. Расширения → Apps Script', None),
        ('  2. Вставьте apps-script/Code.gs', None),
        ('  3. Сохраните, обновите страницу → меню «Транспортный бюджет»', None),
        ('  4. Меняйте Params и жмите «Собрать всё с нуля»', None),
        ('', None),
        ('Интерактивный дашборд без Sheets: dashboard/dashboard.html.', Font(italic=True)),
    ]
    for i, (text, font) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=text)
        if font:
            c.font = font
    ws.column_dimensions['A'].width = 88

    def add(name, rows):
        s = wb.create_sheet(name)
        for r, row in enumerate(rows, 1):
            for ci, v in enumerate(row, 1):
                s.cell(row=r, column=ci, value=v)
        for ci in range(1, len(rows[0]) + 1):
            hc = s.cell(row=1, column=ci)
            hc.font = Font(bold=True, color='FFFFFF')
            hc.fill = PatternFill('solid', fgColor='1F4E79')
        s.freeze_panes = 'A2'
        for ci in range(1, len(rows[0]) + 1):
            ml = max((len(str(rows[r][ci - 1])) for r in range(len(rows))), default=10)
            s.column_dimensions[get_column_letter(ci)].width = min(max(ml + 2, 10), 40)

    add('Vehicles', data['Vehicles'])
    add('Seasonality', data['Seasonality'])
    add('Params', data['Params'])
    # Пример 50 клиентов; полную базу генерирует скрипт
    cust = data['Customers']
    add('Customers', [cust[0]] + cust[1:51])
    add('Бюджет', data['Budget'])
    wb.save(path)


def main():
    if not os.path.exists(INPUTS):
        raise SystemExit('Сначала запустите: node tools/export_inputs.js (создаст tools/out.inputs.json)')
    data = json.load(open(INPUTS, encoding='utf-8'))
    os.makedirs(SAMPLES, exist_ok=True)

    csv_text = build_csv(data)
    with open(os.path.join(SAMPLES, 'transport_budget_data.csv'), 'w', encoding='utf-8') as f:
        f.write(csv_text)
    print('samples/transport_budget_data.csv записан')

    try:
        build_xlsx(data, os.path.join(SAMPLES, 'transport_budget_data.xlsx'))
        print('samples/transport_budget_data.xlsx записан')
    except ImportError:
        print('openpyxl не установлен — xlsx пропущен (pip install openpyxl). CSV записан.')


if __name__ == '__main__':
    main()
