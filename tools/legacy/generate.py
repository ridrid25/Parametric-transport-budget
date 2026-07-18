#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Генератор тестовых данных для параметрической модели транспортного бюджета.
Создаёт transport_budget.xlsx с листами:
  Vehicles, Customers, Seasonality, Params, README
Данные согласованы с логикой Apps Script (Code.gs).
"""
import random
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

random.seed(42)

wb = Workbook()

HEADER_FONT = Font(bold=True)

def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).font = HEADER_FONT
    ws.freeze_panes = "A%d" % (row + 1)

def autosize(ws, widths):
    pass

# ---------------------------------------------------------------------------
# 1. Vehicles (13 машин, 3 весовых категории)
# ---------------------------------------------------------------------------
ws_v = wb.active
ws_v.title = "Vehicles"

veh_header = [
    "vehicle_id", "model", "tier", "capacity_t", "fuel_norm_l_100km",
    "depreciation_month_rub", "insurance_month_rub", "driver_fixed_salary_month_rub",
    "variable_driver_rate_rub_km", "maintenance_rate_rub_km", "max_daily_km"
]
ws_v.append(veh_header)
style_header(ws_v, len(veh_header))

vehicles = [
    # id, model, tier, capacity_t, fuel_norm, deprec, insurance, driver_fixed, driver_var_km, maint_km, max_daily_km
    (1, "ГАЗель NEXT",        "Light",  1.5, 12.5, 16000, 4500, 40000, 3.0, 2.3, 250),
    (2, "ГАЗель NEXT",        "Light",  1.5, 12.5, 16000, 4500, 40000, 3.0, 2.3, 250),
    (3, "Ford Transit",       "Light",  2.0, 13.5, 18000, 4800, 41000, 3.2, 2.5, 250),
    (4, "Ford Transit",       "Light",  2.0, 13.5, 18000, 4800, 41000, 3.2, 2.5, 250),
    (5, "Hyundai HD78",       "Light",  2.5, 15.0, 20000, 5200, 42000, 3.4, 2.7, 260),
    (6, "Isuzu NQR",          "Medium", 5.0, 18.0, 34000, 8500, 48000, 4.0, 3.8, 300),
    (7, "Isuzu NQR",          "Medium", 5.0, 18.0, 34000, 8500, 48000, 4.0, 3.8, 300),
    (8, "Isuzu Forward",      "Medium", 7.0, 20.0, 38000, 9000, 50000, 4.3, 4.1, 300),
    (9, "MAN TGM",            "Medium", 8.0, 22.0, 41000, 9500, 51000, 4.5, 4.3, 320),
    (10, "MAN TGM",           "Medium", 10.0, 24.0, 45000, 10000, 53000, 4.8, 4.6, 320),
    (11, "Volvo FH (тягач)",  "Heavy",  15.0, 28.0, 72000, 14000, 62000, 5.6, 6.0, 550),
    (12, "Scania R (тягач)",  "Heavy",  18.0, 31.0, 82000, 15500, 65000, 6.0, 6.5, 550),
    (13, "Kamaz 65117",       "Heavy",  20.0, 33.0, 90000, 17000, 68000, 6.4, 7.0, 500),
]
for row in vehicles:
    ws_v.append(list(row))

autosize(ws_v, [10, 20, 10, 12, 16, 16, 15, 18, 16, 16, 13])

# ---------------------------------------------------------------------------
# 2. Seasonality (12 месяцев)
# ---------------------------------------------------------------------------
ws_s = wb.create_sheet("Seasonality")
seas_header = ["month_num", "month_name", "demand_index", "fuel_price_index"]
ws_s.append(seas_header)
style_header(ws_s, len(seas_header))

months = [
    (1, "Январь", 0.85, 1.03),
    (2, "Февраль", 0.90, 1.02),
    (3, "Март", 0.95, 1.00),
    (4, "Апрель", 1.00, 0.99),
    (5, "Май", 1.00, 1.00),
    (6, "Июнь", 0.95, 1.01),
    (7, "Июль", 0.90, 1.02),
    (8, "Август", 0.95, 1.02),
    (9, "Сентябрь", 1.05, 1.00),
    (10, "Октябрь", 1.10, 1.00),
    (11, "Ноябрь", 1.25, 1.01),
    (12, "Декабрь", 1.40, 1.04),
]
for row in months:
    ws_s.append(list(row))
autosize(ws_s, [11, 14, 14, 16])

# ---------------------------------------------------------------------------
# 3. Params (ключ-значение)
# ---------------------------------------------------------------------------
ws_p = wb.create_sheet("Params")
par_header = ["param", "value", "description"]
ws_p.append(par_header)
style_header(ws_p, len(par_header))

params = [
    ("fuel_price_rub_per_l", 62, "Базовая цена топлива, руб/л"),
    ("price_per_kg_revenue_default", 220, "Средняя выручка с кг груза (справочно)"),
    ("perdiem_rate_rub_per_night", 2500, "Суточные/командировочные за одну ночёвку в рейсе"),
    ("avg_daily_driving_km", 540, "Средний дневной пробег с учётом норм труда и отдыха (60 км/ч * 9 ч)"),
    ("toll_free_km", 50, "Первые N км от склада без платных дорог"),
    ("toll_rate_rub_per_km", 1.2, "Тариф платных дорог, руб/км сверх toll_free_km"),
    ("outsource_rate_rub_per_km", 45, "Стоимость привлечения стороннего перевозчика (3PL), руб/км, при превышении мощности парка"),
    ("work_days_per_month", 22, "Рабочих дней в месяце"),
    ("load_utilization", 0.85, "Коэффициент использования грузоподъёмности"),
    ("overhead_monthly_fixed_rub", 350000, "Постоянные накладные расходы логистики (диспетчеризация, GPS-мониторинг, ФОТ логистов)"),
    ("revenue_scenario_pct", 0, "Сценарий: изменение выручки, % (меняйте эту ячейку и запускайте \"Пересчитать бюджет\")"),
]
for row in params:
    ws_p.append(list(row))
autosize(ws_p, [32, 14, 70])

# ---------------------------------------------------------------------------
# 4. Customers (~1000 строк)
# ---------------------------------------------------------------------------
ws_c = wb.create_sheet("Customers")
cust_header = [
    "customer_id", "customer_name", "region", "distance_km", "segment",
    "avg_order_weight_kg", "orders_per_month", "price_per_kg_rub",
    "base_monthly_revenue_rub", "tier"
]
ws_c.append(cust_header)
style_header(ws_c, len(cust_header))

regions = [
    "Москва", "Московская обл.", "Санкт-Петербург", "Ленинградская обл.",
    "Казань", "Нижний Новгород", "Екатеринбург", "Самара", "Новосибирск",
    "Ростов-на-Дону", "Краснодар", "Воронеж", "Уфа", "Пермь", "Волгоград",
    "Челябинск", "Красноярск", "Владивосток", "Иркутск", "Омск"
]

def pick_distance(segment):
    # Розница — в основном локальная/региональная доставка (мелкие партии не возят за 2000 км).
    # Опт — региональная/межрегиональная. Крупный дистрибьютор — дальнобойные маршруты
    # (большие партии оправдывают дальние перевозки).
    r = random.random()
    if segment == "Розница":
        if r < 0.75:
            d = random.uniform(5, 100)
        elif r < 0.95:
            d = random.uniform(100, 300)
        else:
            d = random.uniform(300, 600)
    elif segment == "Опт":
        if r < 0.40:
            d = random.uniform(20, 150)
        elif r < 0.80:
            d = random.uniform(150, 500)
        else:
            d = random.uniform(500, 900)
    else:
        if r < 0.30:
            d = random.uniform(100, 400)
        elif r < 0.65:
            d = random.uniform(400, 1000)
        else:
            d = random.uniform(1000, 2000)
    return round(d, 1)

def pick_segment():
    r = random.random()
    if r < 0.70:
        return "Розница"
    elif r < 0.95:
        return "Опт"
    else:
        return "Крупный дистрибьютор"

def segment_profile(segment):
    # (order_weight_range_kg, orders_per_month_range, price_per_kg_range)
    if segment == "Розница":
        return (30, 300), (2, 4), (260, 420)
    elif segment == "Опт":
        return (300, 3000), (1, 3), (180, 280)
    else:
        return (3000, 12000), (1, 2), (120, 190)

def tier_for_weight(w):
    if w < 300:
        return "Light"
    elif w < 3000:
        return "Medium"
    else:
        return "Heavy"

n_customers = 1000
customers = []
for i in range(1, n_customers + 1):
    segment = pick_segment()
    (w_lo, w_hi), (o_lo, o_hi), (p_lo, p_hi) = segment_profile(segment)
    weight = round(random.uniform(w_lo, w_hi), 1)
    orders = random.randint(o_lo, o_hi)
    price = round(random.uniform(p_lo, p_hi), 1)
    distance = pick_distance(segment)
    revenue = round(weight * orders * price, 0)
    region = random.choice(regions)
    tier = tier_for_weight(weight)
    name = f"Клиент {i:04d}"
    customers.append([
        i, name, region, distance, segment, weight, orders, price, revenue, tier
    ])

for row in customers:
    ws_c.append(row)

autosize(ws_c, [12, 34, 18, 12, 10, 18, 15, 15, 20, 10])

# ---------------------------------------------------------------------------
# 5. README
# ---------------------------------------------------------------------------
ws_r = wb.create_sheet("README", 0)
ws_r.column_dimensions["A"].width = 100
lines = [
    "ПАРАМЕТРИЧЕСКИЙ БЮДЖЕТ ТРАНСПОРТНЫХ РАСХОДОВ — тестовые данные",
    "",
    "Состав файла:",
    " • Vehicles — 13 машин (Light/Medium/Heavy), их постоянные и переменные затраты",
    " • Customers — ~1000 покупателей: расстояние от склада (5–2000 км), объём и частота заказов, выручка",
    " • Seasonality — сезонность спроса и цены топлива по месяцам",
    " • Params — параметры модели (цена топлива, тарифы, коэффициенты, СЦЕНАРИЙ изменения выручки)",
    "",
    "Как подключить скрипт:",
    " 1. Откройте это в Google Sheets (файл уже загружен как Google Таблица).",
    " 2. Меню Расширения → Apps Script.",
    " 3. Удалите содержимое Code.gs и вставьте код из файла Code.gs (передан отдельно).",
    " 4. Сохраните проект и обновите страницу таблицы.",
    " 5. В таблице появится меню «Транспортный бюджет»:",
    "      – Пересчитать бюджет — строит помесячный бюджет на листе «Бюджет»",
    "      – Сценарный анализ выручки — считает чувствительность затрат к изменению выручки (лист «Сценарии»)",
    "      – Построить дашборд — создаёт графики на листе «Дашборд»",
    "",
    "Логика модели (кратко):",
    " • Выручка каждого клиента масштабируется сезонным индексом месяца и сценарием revenue_scenario_pct (лист Params).",
    " • Объём груза = выручка / цена за кг → количество рейсов = объём / (грузоподъёмность машины его тарифной группы × загрузка).",
    " • Затраты на рейс: топливо, ТО, платные дороги, сдельная доплата водителю, суточные (если рейс требует ночёвки).",
    " • Если требуемый пробег превышает мощность собственного парка (13 машин, лимит км/день), излишек",
    "   передаётся на аутсорс (3PL) по более высокой ставке — это создаёт нелинейность расходов при росте выручки.",
    " • Постоянные расходы (амортизация, страховка, оклады, накладные) не зависят от выручки.",
    " • Итог: доля переменных затрат растёт почти пропорционально выручке, доля постоянных — падает (эффект масштаба),",
    "   но при исчерпании мощности парка предельная стоимость км резко возрастает (эффект аутсорса).",
]
for i, l in enumerate(lines, start=1):
    ws_r.cell(row=i, column=1, value=l)
    if i == 1:
        ws_r.cell(row=i, column=1).font = Font(bold=True, size=14)

out_path = "/tmp/claude-0/-home-user-Test-claud-parametr/cb645b49-a197-51ff-910a-b3a035210890/scratchpad/transport_budget/transport_budget.xlsx"
wb.save(out_path)
print("saved:", out_path)
print("customers:", len(customers))
print("total base annual revenue (rub):", sum(c[8] for c in customers) * 12)
